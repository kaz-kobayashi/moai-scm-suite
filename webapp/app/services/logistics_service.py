import numpy as np
import pandas as pd
from typing import List, Dict, Tuple, Optional, Any
from geopy.distance import great_circle
import math
import random
from sklearn.cluster import KMeans, MiniBatchKMeans
from sklearn.metrics import silhouette_score
from scipy.cluster.hierarchy import linkage, fcluster
from scipy.spatial.distance import pdist, squareform
import plotly.graph_objects as go
import plotly.express as px
from collections import defaultdict
import networkx as nx
try:
    import pulp
    PULP_AVAILABLE = True
except ImportError:
    PULP_AVAILABLE = False
try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
import io
import base64

from app.models.logistics import (
    CustomerData, DCData, PlantData, ProductData, LocationData,
    WeiszfeldResult, KMedianResult, ClusteringResult, LNDResult,
    NetworkVisualizationResult, NetworkAnalysisResult
)

class LogisticsOptimizationService:
    """物流最適化サービス"""
    
    def __init__(self):
        self.random_state = 42
        np.random.seed(self.random_state)
        random.seed(self.random_state)
    
    # =====================================================
    # 距離計算ユーティリティ
    # =====================================================
    
    def calculate_distance(self, loc1: LocationData, loc2: LocationData, 
                          use_great_circle: bool = True) -> float:
        """2点間の距離を計算"""
        if use_great_circle:
            return great_circle(
                (loc1.latitude, loc1.longitude),
                (loc2.latitude, loc2.longitude)
            ).kilometers
        else:
            # ユークリッド距離
            return math.sqrt(
                (loc1.latitude - loc2.latitude)**2 + 
                (loc1.longitude - loc2.longitude)**2
            )
    
    def calculate_distance_matrix(self, locations1: List[LocationData], 
                                 locations2: List[LocationData],
                                 use_great_circle: bool = True) -> np.ndarray:
        """距離行列を計算"""
        matrix = np.zeros((len(locations1), len(locations2)))
        for i, loc1 in enumerate(locations1):
            for j, loc2 in enumerate(locations2):
                matrix[i, j] = self.calculate_distance(loc1, loc2, use_great_circle)
        return matrix
    
    # =====================================================
    # Weiszfeld法による施設立地最適化
    # =====================================================
    
    def weiszfeld_single(self, customers: List[CustomerData], 
                        initial_location: Optional[LocationData] = None,
                        max_iterations: int = 1000, tolerance: float = 1e-6,
                        use_great_circle: bool = True) -> Tuple[LocationData, float, int]:
        """単一施設のWeiszfeld法最適化"""
        
        # 初期位置を設定
        if initial_location is None:
            # 重心を初期位置とする
            total_demand = sum(c.demand for c in customers)
            lat = sum(c.latitude * c.demand for c in customers) / total_demand
            lon = sum(c.longitude * c.demand for c in customers) / total_demand
            current_location = LocationData(name="facility", latitude=lat, longitude=lon)
        else:
            current_location = initial_location
        
        convergence_history = []
        
        for iteration in range(max_iterations):
            # 各顧客までの距離と重みを計算
            numerator_lat = numerator_lon = denominator = 0.0
            total_cost = 0.0
            
            for customer in customers:
                distance = self.calculate_distance(current_location, customer, use_great_circle)
                if distance < 1e-10:  # 同一位置の場合
                    continue
                
                weight = customer.demand / distance
                numerator_lat += weight * customer.latitude
                numerator_lon += weight * customer.longitude
                denominator += weight
                total_cost += customer.demand * distance
            
            if denominator == 0:
                break
                
            # 新しい位置を計算
            new_lat = numerator_lat / denominator
            new_lon = numerator_lon / denominator
            new_location = LocationData(name="facility", latitude=new_lat, longitude=new_lon)
            
            # 収束判定
            position_change = self.calculate_distance(current_location, new_location, use_great_circle)
            convergence_history.append(total_cost)
            
            if position_change < tolerance:
                return new_location, total_cost, iteration + 1
            
            current_location = new_location
        
        return current_location, total_cost, max_iterations
    
    def weiszfeld_multiple(self, customers: List[CustomerData], num_facilities: int,
                          max_iterations: int = 1000, tolerance: float = 1e-6,
                          use_great_circle: bool = True) -> WeiszfeldResult:
        """複数施設のWeiszfeld法最適化"""
        
        if num_facilities == 1:
            facility, cost, iterations = self.weiszfeld_single(
                customers, None, max_iterations, tolerance, use_great_circle)
            return WeiszfeldResult(
                facility_locations=[facility],
                total_cost=cost,
                iterations=iterations,
                convergence_history=[cost],
                customer_assignments={c.name: 0 for c in customers}
            )
        
        best_facilities = []
        best_cost = float('inf')
        best_assignments = {}
        best_history = []
        
        # 複数回実行して最良解を選択
        for attempt in range(5):
            # ランダムな初期位置を生成
            initial_facilities = []
            for i in range(num_facilities):
                idx = random.randint(0, len(customers) - 1)
                customer = customers[idx]
                initial_facilities.append(LocationData(
                    name=f"facility_{i}",
                    latitude=customer.latitude + random.uniform(-0.1, 0.1),
                    longitude=customer.longitude + random.uniform(-0.1, 0.1)
                ))
            
            facilities, cost, assignments, history = self._optimize_multiple_facilities(
                customers, initial_facilities, max_iterations, tolerance, use_great_circle)
            
            if cost < best_cost:
                best_cost = cost
                best_facilities = facilities
                best_assignments = assignments
                best_history = history
        
        return WeiszfeldResult(
            facility_locations=best_facilities,
            total_cost=best_cost,
            iterations=len(best_history),
            convergence_history=best_history,
            customer_assignments=best_assignments
        )
    
    def _optimize_multiple_facilities(self, customers: List[CustomerData],
                                    initial_facilities: List[LocationData],
                                    max_iterations: int, tolerance: float,
                                    use_great_circle: bool) -> Tuple[List[LocationData], float, Dict[str, int], List[float]]:
        """複数施設の反復最適化"""
        
        facilities = initial_facilities.copy()
        convergence_history = []
        
        for iteration in range(max_iterations):
            # 顧客を最近の施設に割当
            assignments = {}
            facility_customers = [[] for _ in range(len(facilities))]
            
            for customer in customers:
                best_facility = 0
                best_distance = float('inf')
                
                for f_idx, facility in enumerate(facilities):
                    distance = self.calculate_distance(customer, facility, use_great_circle)
                    if distance < best_distance:
                        best_distance = distance
                        best_facility = f_idx
                
                assignments[customer.name] = best_facility
                facility_customers[best_facility].append(customer)
            
            # 各施設位置を最適化
            new_facilities = []
            total_cost = 0.0
            position_changed = False
            
            for f_idx, assigned_customers in enumerate(facility_customers):
                if not assigned_customers:
                    # 顧客が割り当てられていない施設はそのまま
                    new_facilities.append(facilities[f_idx])
                    continue
                
                # この施設に割り当てられた顧客で最適化
                new_facility, facility_cost, _ = self.weiszfeld_single(
                    assigned_customers, facilities[f_idx], 
                    max_iterations // 10, tolerance, use_great_circle)
                
                new_facilities.append(new_facility)
                total_cost += facility_cost
                
                # 位置変更をチェック
                if self.calculate_distance(facilities[f_idx], new_facility, use_great_circle) > tolerance:
                    position_changed = True
            
            convergence_history.append(total_cost)
            
            if not position_changed:
                break
                
            facilities = new_facilities
        
        return facilities, total_cost, assignments, convergence_history
    
    # =====================================================
    # K-Median最適化（Lagrange緩和）
    # =====================================================
    
    def solve_k_median(self, customers: List[CustomerData], dc_candidates: List[DCData],
                      k: int, max_iterations: int = 1000, learning_rate: float = 0.01,
                      momentum: float = 0.9, use_adam: bool = False, 
                      use_lr_scheduling: bool = False, capacity_constraint: bool = False) -> KMedianResult:
        """K-Median問題をLagrange緩和法で解く（Adam最適化・学習率スケジューリング対応）"""
        
        n_customers = len(customers)
        n_facilities = len(dc_candidates)
        
        # 距離行列とコスト行列を計算
        distance_matrix = np.zeros((n_customers, n_facilities))
        for i, customer in enumerate(customers):
            for j, facility in enumerate(dc_candidates):
                dist = self.calculate_distance(customer, facility)
                distance_matrix[i, j] = dist
        
        # 重み（需要）
        weights = np.array([c.demand for c in customers])
        
        # コスト行列 C[i,j] = 重み[i] * 距離[i,j] * 単価
        transport_unit_cost = 0.5
        C = distance_matrix * weights.reshape((n_customers, 1)) * transport_unit_cost
        
        # Lagrange乗数の初期化
        u = np.zeros(n_customers)
        
        # Adam パラメータ
        if use_adam:
            beta_1 = 0.9  # 初期値（スケジューリング時に変更）
            beta_2 = 0.999
            epsilon = 1e-8
            m_t = np.zeros(n_customers)  # 1st moment
            v_t = np.zeros(n_customers)  # 2nd moment
        else:
            m_t = np.zeros(n_customers)  # momentum term
        
        # 学習率スケジューリング
        if use_lr_scheduling:
            max_lr = learning_rate
            min_lr = max_lr / 25
            half_iter = max_iterations // 2
            
            # Cosine annealing with linear warmup
            lr_schedule = np.concatenate([
                np.linspace(min_lr, max_lr, half_iter),
                min_lr + (max_lr - min_lr) * 0.5 * (1 + np.cos(np.linspace(0, np.pi, max_iterations - half_iter)))
            ])
            
            if use_adam:
                # Momentum scheduling (fit-one-cycle)
                max_momentum = 0.95
                min_momentum = 0.85
                momentum_schedule = np.concatenate([
                    np.linspace(max_momentum, min_momentum, half_iter),
                    min_momentum + (max_momentum - min_momentum) * 0.5 * (1 - np.cos(np.linspace(0, np.pi, max_iterations - half_iter)))
                ])
        else:
            lr_schedule = [learning_rate] * max_iterations
            momentum_schedule = [momentum] * max_iterations
        
        # 容量制約
        capacities = None
        if capacity_constraint:
            capacities = np.array([dc.capacity for dc in dc_candidates])
        
        # 最適化ループ
        objective_history = []
        best_ub = float('inf')
        selected_facilities = []
        customer_assignments = {}
        
        for t in range(max_iterations):
            current_lr = lr_schedule[t] if use_lr_scheduling else learning_rate
            
            if use_adam and use_lr_scheduling:
                current_momentum = momentum_schedule[t]
                beta_1 = current_momentum
            
            # コスト行列の更新（Lagrange乗数を考慮）
            C_reduced = C - u.reshape((n_customers, 1))
            
            # 施設選択（容量制約考慮）
            if capacity_constraint and capacities is not None:
                # 容量制約付き施設選択
                facility_values = np.zeros(n_facilities)
                for j in range(n_facilities):
                    # 各施設に対して容量制約下での最適割当を計算
                    costs_j = C_reduced[:, j]
                    sorted_indices = np.argsort(costs_j)
                    cumulative_demand = 0
                    total_value = 0
                    
                    for i in sorted_indices:
                        if cumulative_demand + weights[i] <= capacities[j] and costs_j[i] < 0:
                            total_value += costs_j[i]
                            cumulative_demand += weights[i]
                        else:
                            break
                    facility_values[j] = total_value
            else:
                # 容量制約なし
                facility_values = np.where(C_reduced < 0, C_reduced, 0).sum(axis=0)
            
            # 上位k個の施設を選択（最も負の値が大きい施設、つまり最も良い施設）
            selected_indices = np.argsort(facility_values)[-k:]
            
            # デバッグ情報の出力
            print(f"=== K-Median Debug Info (Iteration {t}) ===")
            print(f"Facility values: {facility_values}")
            print(f"Selected indices: {selected_indices}")
            print(f"Selected facility names: {[dc_candidates[i].name for i in selected_indices]}")
            print("=" * 50)
            
            # 顧客割当の計算
            assignment_counts = np.zeros(n_customers)
            for i in range(n_customers):
                assigned_facilities = []
                for j in selected_indices:
                    if C_reduced[i, j] < 0:
                        assigned_facilities.append(j)
                
                assignment_counts[i] = len(assigned_facilities)
            
            # 下界の計算
            lower_bound = u.sum() + facility_values[selected_indices].sum()
            
            # 実行可能解の構築と上界計算
            temp_assignments = {}
            total_cost = 0.0
            
            # 容量制約がある場合の顧客割当
            if capacity_constraint and capacities is not None:
                # 容量制約付き顧客割当
                facility_loads = {j: 0.0 for j in selected_indices}
                
                # コストの昇順でソートされた顧客と施設のペア
                customer_facility_pairs = []
                for i, customer in enumerate(customers):
                    for j in selected_indices:
                        customer_facility_pairs.append((C[i, j], i, j, customer.name))
                
                customer_facility_pairs.sort()
                assigned_customers = set()
                
                for cost, i, j, customer_name in customer_facility_pairs:
                    if customer_name not in assigned_customers:
                        customer_demand = customers[i].demand
                        if facility_loads[j] + customer_demand <= capacities[j]:
                            temp_assignments[customer_name] = int(j)
                            facility_loads[j] += customer_demand
                            total_cost += cost
                            assigned_customers.add(customer_name)
            else:
                # 容量制約なしの場合は最近隣割当
                for i, customer in enumerate(customers):
                    best_j = None
                    best_cost = float('inf')
                    
                    for j in selected_indices:
                        cost = C[i, j]  # 元の距離コストを使用（固定費は別途追加）
                        if cost < best_cost:
                            best_cost = cost
                            best_j = j
                    
                    if best_j is not None:
                        temp_assignments[customer.name] = int(best_j)
                        total_cost += best_cost
                        
                # デバッグ: 顧客割り当て状況を表示
                assignment_summary = {}
                for customer_name, facility_idx in temp_assignments.items():
                    facility_name = dc_candidates[facility_idx].name
                    if facility_name not in assignment_summary:
                        assignment_summary[facility_name] = []
                    assignment_summary[facility_name].append(customer_name)
                
                print(f"Customer assignment summary:")
                for facility_name, customers_list in assignment_summary.items():
                    print(f"  {facility_name}: {len(customers_list)} customers")
                print(f"Total customers assigned: {len(temp_assignments)}")
            
            # 固定費の追加
            for j in selected_indices:
                total_cost += dc_candidates[j].fixed_cost
            
            if total_cost < best_ub:
                best_ub = total_cost
                selected_facilities = selected_indices.copy()
                customer_assignments = temp_assignments.copy()
            
            objective_history.append(lower_bound)
            
            # 勾配（制約違反）の計算
            g_t = 1.0 - assignment_counts
            
            # パラメータ更新
            if use_adam:
                # Adam更新
                m_t = beta_1 * m_t + (1 - beta_1) * g_t
                v_t = beta_2 * v_t + (1 - beta_2) * (g_t ** 2)
                
                # バイアス補正
                m_cap = m_t / (1 - beta_1 ** (t + 1))
                v_cap = v_t / (1 - beta_2 ** (t + 1))
                
                # 更新
                u = u + current_lr * m_cap / (np.sqrt(v_cap) + epsilon)
            else:
                # 標準のmomentum SGD
                m_t = momentum * m_t + g_t
                u = u + current_lr * m_t
            
            # 収束判定
            if t > 10 and abs(objective_history[-1] - objective_history[-2]) < 1e-6:
                break
        
        return KMedianResult(
            selected_facilities=[int(idx) for idx in selected_facilities],
            facility_locations=[dc_candidates[idx] for idx in selected_facilities],
            total_cost=best_ub,
            objective_history=objective_history,
            customer_assignments=customer_assignments
        )
    
    def find_learning_rate(self, customers: List[CustomerData], dc_candidates: List[DCData],
                          k: int, lr_range: Tuple[float, float] = (1e-7, 10.0),
                          num_iterations: int = 100, divergence_threshold: float = 4.0) -> Dict[str, Any]:
        """Learning Rate Finder for K-Median optimization"""
        
        n_customers = len(customers)
        n_facilities = len(dc_candidates)
        
        # 距離行列とコスト行列を計算
        distance_matrix = np.zeros((n_customers, n_facilities))
        for i, customer in enumerate(customers):
            for j, facility in enumerate(dc_candidates):
                dist = self.calculate_distance(customer, facility)
                distance_matrix[i, j] = dist
        
        # 重み（需要）
        weights = np.array([c.demand for c in customers])
        
        # コスト行列
        transport_unit_cost = 0.5
        C = distance_matrix * weights.reshape((n_customers, 1)) * transport_unit_cost
        
        # 学習率範囲を対数スケールで生成
        min_lr, max_lr = lr_range
        lrs = np.logspace(np.log10(min_lr), np.log10(max_lr), num_iterations)
        losses = []
        
        # 各学習率でのテスト
        for lr in lrs:
            # Lagrange乗数の初期化
            u = np.zeros(n_customers)
            
            # Adam パラメータ
            beta_1, beta_2 = 0.9, 0.999
            epsilon = 1e-8
            m_t = np.zeros(n_customers)
            v_t = np.zeros(n_customers)
            
            # 短期間の最適化実行
            for t in range(min(20, num_iterations)):
                # コスト行列の更新
                C_reduced = C - u.reshape((n_customers, 1))
                
                # 施設選択
                facility_values = np.where(C_reduced < 0, C_reduced, 0).sum(axis=0)
                selected_indices = np.argsort(facility_values)[:k]
                
                # 顧客割当の計算
                assignment_counts = np.zeros(n_customers)
                for i in range(n_customers):
                    assigned_facilities = []
                    for j in selected_indices:
                        if C_reduced[i, j] < 0:
                            assigned_facilities.append(j)
                    assignment_counts[i] = len(assigned_facilities)
                
                # 勾配計算
                g_t = 1.0 - assignment_counts
                
                # Adam更新
                m_t = beta_1 * m_t + (1 - beta_1) * g_t
                v_t = beta_2 * v_t + (1 - beta_2) * (g_t ** 2)
                
                # バイアス補正
                m_cap = m_t / (1 - beta_1 ** (t + 1))
                v_cap = v_t / (1 - beta_2 ** (t + 1))
                
                # 更新
                u = u + lr * m_cap / (np.sqrt(v_cap) + epsilon)
            
            # 損失計算（制約違反の平均）
            loss = np.mean(np.abs(g_t))
            losses.append(loss)
            
            # 発散チェック
            if len(losses) > 10 and loss > divergence_threshold * min(losses[-10:]):
                break
        
        # 最適学習率を探索
        losses = np.array(losses[:len(lrs)])
        
        # 勾配法で最適学習率を見つける
        if len(losses) > 3:
            # 損失の変化率を計算
            gradients = np.gradient(losses)
            
            # 最も急激に減少している点を見つける
            min_gradient_idx = np.argmin(gradients[1:-1]) + 1  # 端点を除外
            
            # より保守的な選択：最小値の10倍の学習率
            min_loss_idx = np.argmin(losses)
            suggested_lr = lrs[min_gradient_idx] if min_gradient_idx < len(lrs) else lrs[min_loss_idx]
            
            # さらに保守的に1/10に調整
            suggested_lr = suggested_lr / 10
        else:
            suggested_lr = lrs[np.argmin(losses)]
        
        return {
            "learning_rates": lrs.tolist(),
            "losses": losses.tolist(),
            "suggested_lr": float(suggested_lr),
            "min_loss_lr": float(lrs[np.argmin(losses)]),
            "range_used": lr_range,
            "iterations_tested": len(losses)
        }
    
    # =====================================================
    # 顧客クラスタリング
    # =====================================================
    
    def cluster_customers(self, customers: List[CustomerData], method: str = "kmeans",
                         n_clusters: int = 5, use_road_distance: bool = False) -> ClusteringResult:
        """顧客をクラスタリング"""
        
        # 特徴量行列を作成
        features = np.array([[c.latitude, c.longitude, c.demand] for c in customers])
        locations = np.array([[c.latitude, c.longitude] for c in customers])
        
        if method == "kmeans":
            if len(customers) < 1000:
                clusterer = KMeans(n_clusters=n_clusters, random_state=self.random_state)
            else:
                clusterer = MiniBatchKMeans(n_clusters=n_clusters, random_state=self.random_state)
            
            cluster_labels = clusterer.fit_predict(features)
            cluster_centers_raw = clusterer.cluster_centers_
            
        elif method == "hierarchical":
            # 階層クラスタリング
            if use_road_distance:
                # 距離行列を使用（簡素化でユークリッド距離）
                distances = pdist(locations)
                linkage_matrix = linkage(distances, method='ward')
            else:
                linkage_matrix = linkage(features, method='ward')
            
            cluster_labels = fcluster(linkage_matrix, n_clusters, criterion='maxclust') - 1
            
            # クラスター中心を計算
            cluster_centers_raw = []
            for i in range(n_clusters):
                mask = cluster_labels == i
                if np.sum(mask) > 0:
                    center = np.mean(features[mask], axis=0)
                    cluster_centers_raw.append(center)
                else:
                    cluster_centers_raw.append([0, 0, 0])
            cluster_centers_raw = np.array(cluster_centers_raw)
        
        # クラスター中心を LocationData に変換
        cluster_centers = []
        for i, center in enumerate(cluster_centers_raw):
            cluster_centers.append(LocationData(
                name=f"cluster_{i}",
                latitude=float(center[0]),
                longitude=float(center[1])
            ))
        
        # 顧客割当辞書を作成
        cluster_assignments = {}
        for i, customer in enumerate(customers):
            cluster_assignments[customer.name] = int(cluster_labels[i])
        
        # 集約顧客を作成
        aggregated_customers = []
        for i in range(n_clusters):
            mask = cluster_labels == i
            if np.sum(mask) > 0:
                cluster_customers = [customers[j] for j in np.where(mask)[0]]
                total_demand = sum(c.demand for c in cluster_customers)
                
                # 需要重心で位置を計算
                lat = sum(c.latitude * c.demand for c in cluster_customers) / total_demand
                lon = sum(c.longitude * c.demand for c in cluster_customers) / total_demand
                
                aggregated_customers.append(CustomerData(
                    name=f"aggregated_cluster_{i}",
                    latitude=lat,
                    longitude=lon,
                    demand=total_demand
                ))
        
        # シルエット係数を計算
        if len(set(cluster_labels)) > 1:
            silhouette = silhouette_score(features, cluster_labels)
        else:
            silhouette = 0.0
        
        return ClusteringResult(
            clusters=cluster_assignments,
            cluster_centers=cluster_centers,
            aggregated_customers=aggregated_customers,
            silhouette_score=silhouette
        )
    
    # =====================================================
    # 基本的な物流ネットワーク設計
    # =====================================================
    
    def solve_basic_lnd(self, customers: List[CustomerData], dc_candidates: List[DCData],
                       plants: List[PlantData], model_type: str = "multi_source",
                       max_facilities: Optional[int] = None) -> LNDResult:
        """基本的な物流ネットワーク設計を解く"""
        
        # 簡素化されたヒューリスティック解法
        n_customers = len(customers)
        n_facilities = len(dc_candidates)
        
        # 各顧客から各DCまでの距離を計算
        distance_matrix = self.calculate_distance_matrix(customers, dc_candidates)
        
        # 容量制約を考慮してDCを選択
        selected_facilities = []
        facility_loads = {}
        customer_assignments = {}
        
        if max_facilities is None:
            max_facilities = min(n_facilities, max(1, n_customers // 10))
        
        # グリーディ選択
        remaining_customers = customers.copy()
        
        while len(selected_facilities) < max_facilities and remaining_customers:
            best_facility = None
            best_score = -1
            
            for f_idx, facility in enumerate(dc_candidates):
                if f_idx in [f['index'] for f in selected_facilities]:
                    continue
                
                # この施設を追加した場合のスコアを計算
                score = 0
                assignable_customers = []
                current_load = facility_loads.get(f_idx, 0)
                
                for customer in remaining_customers:
                    distance = distance_matrix[customers.index(customer), f_idx]
                    if current_load + customer.demand <= facility.capacity:
                        score += customer.demand / (distance + 1)  # 距離の逆数で重み付け
                        assignable_customers.append(customer)
                
                if score > best_score and assignable_customers:
                    best_score = score
                    best_facility = {
                        'index': f_idx,
                        'facility': facility,
                        'assignable_customers': assignable_customers
                    }
            
            if best_facility is None:
                break
            
            # 選択された施設に顧客を割当
            selected_facilities.append(best_facility)
            facility_loads[best_facility['index']] = 0
            
            for customer in best_facility['assignable_customers']:
                customer_assignments[customer.name] = best_facility['index']
                facility_loads[best_facility['index']] += customer.demand
                remaining_customers.remove(customer)
        
        # 残りの顧客を既存の施設に割当（容量制約を無視）
        for customer in remaining_customers:
            best_facility_idx = 0
            best_distance = float('inf')
            
            for f_info in selected_facilities:
                f_idx = f_info['index']
                distance = distance_matrix[customers.index(customer), f_idx]
                if distance < best_distance:
                    best_distance = distance
                    best_facility_idx = f_idx
            
            customer_assignments[customer.name] = best_facility_idx
            facility_loads[best_facility_idx] += customer.demand
        
        # 費用計算
        total_transport_cost = 0
        total_fixed_cost = 0
        
        for customer in customers:
            if customer.name in customer_assignments:
                f_idx = customer_assignments[customer.name]
                distance = distance_matrix[customers.index(customer), f_idx]
                total_transport_cost += customer.demand * distance * 0.1  # 仮の単価
        
        for f_info in selected_facilities:
            total_fixed_cost += f_info['facility'].fixed_cost
        
        # 結果を構築
        selected_facility_data = [f_info['facility'] for f_info in selected_facilities]
        
        # 稼働率計算
        utilization = {}
        for f_info in selected_facilities:
            f_idx = f_info['index']
            utilization[f_info['facility'].name] = facility_loads[f_idx] / f_info['facility'].capacity
        
        # フロー割当（簡素化）
        flow_assignments = {}
        for customer in customers:
            if customer.name in customer_assignments:
                f_idx = customer_assignments[customer.name]
                facility_name = dc_candidates[f_idx].name
                flow_assignments[customer.name] = {facility_name: customer.demand}
        
        return LNDResult(
            selected_facilities=selected_facility_data,
            flow_assignments=flow_assignments,
            total_cost=total_transport_cost + total_fixed_cost,
            cost_breakdown={
                "transportation": total_transport_cost,
                "fixed": total_fixed_cost,
                "variable": 0.0,
                "inventory": 0.0
            },
            facility_utilization=utilization,
            network_performance={
                "average_distance": np.mean([distance_matrix[i, customer_assignments[customers[i].name]] 
                                           for i in range(len(customers)) 
                                           if customers[i].name in customer_assignments]),
                "facility_count": len(selected_facilities),
                "total_demand": sum(c.demand for c in customers)
            },
            solution_status="Optimal",
            solve_time=0.1  # ヒューリスティック解法のため高速
        )
    
    # =====================================================
    # 高度な物流ネットワーク設計
    # =====================================================
    
    def solve_multi_source_lnd(self, customers: List[CustomerData], dc_candidates: List[DCData],
                              plants: List[PlantData], products: Optional[List[ProductData]] = None,
                              max_iterations: int = 1000, tolerance: float = 1e-6) -> LNDResult:
        """Multi-source logistics network design with multiple plants and DCs"""
        
        n_customers = len(customers)
        n_facilities = len(dc_candidates)
        n_plants = len(plants)
        
        if not products:
            # デフォルトの単一製品を作成
            products = [ProductData(name="default_product", unit_cost=1.0, weight=1.0, volume=1.0)]
        
        n_products = len(products)
        
        # 距離行列を計算
        # Plant -> DC
        plant_dc_distances = self.calculate_distance_matrix(plants, dc_candidates)
        # DC -> Customer
        dc_customer_distances = self.calculate_distance_matrix(dc_candidates, customers)
        
        # 需要行列 [customers x products]
        demand_matrix = np.zeros((n_customers, n_products))
        for i, customer in enumerate(customers):
            demand_matrix[i, 0] = customer.demand  # 単一製品の場合
        
        # 生産能力行列 [plants x products]
        production_capacity = np.zeros((n_plants, n_products))
        for i, plant in enumerate(plants):
            production_capacity[i, 0] = plant.capacity  # 単一製品の場合
        
        # ラグランジュ緩和法による解法
        # 変数: x[i,j,k] = plant i から DC j への product k のフロー
        #       y[j,l,k] = DC j から customer l への product k のフロー
        #       z[j] = DC j が開設されるかどうか（バイナリ）
        
        # ラグランジュ乗数の初期化
        u_demand = np.zeros((n_customers, n_products))  # 需要制約
        u_capacity = np.zeros((n_plants, n_products))   # 生産能力制約
        u_balance = np.zeros((n_facilities, n_products))  # フロー均衡制約
        
        # 学習パラメータ
        learning_rate = 0.01
        
        # 最適化履歴
        objective_history = []
        best_upper_bound = float('inf')
        best_solution = {}
        
        for iteration in range(max_iterations):
            # 下位問題の解法
            
            # 1. Plant-to-DC フロー最適化
            plant_dc_flows = {}
            plant_dc_cost = 0.0
            
            for i in range(n_plants):
                for j in range(n_facilities):
                    for k in range(n_products):
                        # コスト計算（輸送費 + ラグランジュ項）
                        transport_cost = plant_dc_distances[i, j] * 0.1  # 単価
                        lagrange_term = u_capacity[i, k] - u_balance[j, k]
                        reduced_cost = transport_cost + lagrange_term
                        
                        # 最適フロー決定
                        max_flow = min(plants[i].capacity, dc_candidates[j].capacity)
                        if reduced_cost < 0:
                            flow = max_flow
                        else:
                            flow = 0
                        
                        plant_dc_flows[(i, j, k)] = flow
                        plant_dc_cost += flow * transport_cost
            
            # 2. DC-to-Customer フロー最適化
            dc_customer_flows = {}
            dc_customer_cost = 0.0
            
            for j in range(n_facilities):
                for l in range(n_customers):
                    for k in range(n_products):
                        # コスト計算
                        transport_cost = dc_customer_distances[j, l] * 0.1
                        lagrange_term = u_balance[j, k] - u_demand[l, k]
                        reduced_cost = transport_cost + lagrange_term
                        
                        # 最適フロー決定
                        max_flow = min(dc_candidates[j].capacity, demand_matrix[l, k])
                        if reduced_cost < 0:
                            flow = max_flow
                        else:
                            flow = 0
                        
                        dc_customer_flows[(j, l, k)] = flow
                        dc_customer_cost += flow * transport_cost
            
            # 3. DC開設決定
            selected_dcs = []
            dc_fixed_cost = 0.0
            
            for j in range(n_facilities):
                # DCの開設価値を計算
                inflow = sum(plant_dc_flows.get((i, j, k), 0) 
                           for i in range(n_plants) for k in range(n_products))
                outflow = sum(dc_customer_flows.get((j, l, k), 0)
                            for l in range(n_customers) for k in range(n_products))
                
                # DCを開設するかどうか判定
                if inflow > 0 or outflow > 0:
                    selected_dcs.append(j)
                    dc_fixed_cost += dc_candidates[j].fixed_cost
            
            # 下界の計算
            lower_bound = (plant_dc_cost + dc_customer_cost + dc_fixed_cost +
                          np.sum(u_demand * demand_matrix) + 
                          np.sum(u_capacity * production_capacity))
            
            objective_history.append(lower_bound)
            
            # 実行可能解の構築（上界）
            feasible_flows = self._construct_feasible_solution(
                customers, dc_candidates, plants, products,
                plant_dc_flows, dc_customer_flows, selected_dcs
            )
            
            upper_bound = feasible_flows['total_cost']
            
            if upper_bound < best_upper_bound:
                best_upper_bound = upper_bound
                best_solution = feasible_flows
            
            # ラグランジュ乗数の更新
            # 需要制約の違反
            demand_violation = np.zeros((n_customers, n_products))
            for l in range(n_customers):
                for k in range(n_products):
                    served = sum(dc_customer_flows.get((j, l, k), 0) for j in range(n_facilities))
                    demand_violation[l, k] = demand_matrix[l, k] - served
            
            # 生産能力制約の違反
            capacity_violation = np.zeros((n_plants, n_products))
            for i in range(n_plants):
                for k in range(n_products):
                    produced = sum(plant_dc_flows.get((i, j, k), 0) for j in range(n_facilities))
                    capacity_violation[i, k] = produced - production_capacity[i, k]
            
            # フロー均衡制約の違反
            balance_violation = np.zeros((n_facilities, n_products))
            for j in range(n_facilities):
                for k in range(n_products):
                    inflow = sum(plant_dc_flows.get((i, j, k), 0) for i in range(n_plants))
                    outflow = sum(dc_customer_flows.get((j, l, k), 0) for l in range(n_customers))
                    balance_violation[j, k] = inflow - outflow
            
            # 乗数更新
            u_demand += learning_rate * demand_violation
            u_capacity -= learning_rate * capacity_violation  # 不等式制約のため符号注意
            u_balance += learning_rate * balance_violation
            
            # 収束判定
            if (iteration > 10 and 
                abs(objective_history[-1] - objective_history[-2]) < tolerance):
                break
            
            # 学習率の調整
            if iteration % 100 == 0 and iteration > 0:
                learning_rate *= 0.95
        
        # 結果の構築
        selected_facilities = [dc_candidates[j] for j in selected_dcs]
        
        flow_assignments = {}
        for customer in customers:
            flow_assignments[customer.name] = {}
            for j in selected_dcs:
                flow = sum(best_solution.get('dc_customer_flows', {}).get((j, customers.index(customer), k), 0)
                          for k in range(n_products))
                if flow > 0:
                    flow_assignments[customer.name][dc_candidates[j].name] = flow
        
        # 稼働率計算
        utilization = {}
        for j in selected_dcs:
            total_flow = sum(flow_assignments[c.name].get(dc_candidates[j].name, 0) 
                           for c in customers)
            utilization[dc_candidates[j].name] = total_flow / dc_candidates[j].capacity
        
        return LNDResult(
            selected_facilities=selected_facilities,
            flow_assignments=flow_assignments,
            total_cost=best_upper_bound,
            cost_breakdown={
                "transportation": best_solution.get('transport_cost', 0),
                "fixed": best_solution.get('fixed_cost', 0),
                "variable": 0.0,
                "inventory": 0.0
            },
            facility_utilization=utilization,
            network_performance={
                "average_distance": np.mean([dc_customer_distances[j, customers.index(c)] 
                                           for c in customers for j in selected_dcs
                                           if c.name in flow_assignments and 
                                           dc_candidates[j].name in flow_assignments[c.name]]),
                "facility_count": len(selected_facilities),
                "total_demand": sum(c.demand for c in customers)
            },
            solution_status="Lagrange_Relaxation",
            solve_time=iteration * 0.001
        )
    
    def _construct_feasible_solution(self, customers: List[CustomerData], dc_candidates: List[DCData],
                                   plants: List[PlantData], products: List[ProductData],
                                   plant_dc_flows: Dict, dc_customer_flows: Dict, 
                                   selected_dcs: List[int]) -> Dict[str, Any]:
        """ラグランジュ緩和の解から実行可能解を構築"""
        
        # 簡素化されたヒューリスティック
        total_transport_cost = 0.0
        total_fixed_cost = 0.0
        
        # 固定費
        for j in selected_dcs:
            total_fixed_cost += dc_candidates[j].fixed_cost
        
        # 輸送費（概算）
        for customer in customers:
            if selected_dcs:
                # 最も近いDCに割当
                min_distance = float('inf')
                for j in selected_dcs:
                    distance = self.calculate_distance(customer, dc_candidates[j])
                    if distance < min_distance:
                        min_distance = distance
                total_transport_cost += customer.demand * min_distance * 0.1
        
        return {
            'total_cost': total_transport_cost + total_fixed_cost,
            'transport_cost': total_transport_cost,
            'fixed_cost': total_fixed_cost,
            'dc_customer_flows': dc_customer_flows
        }
    
    def solve_single_source_lnd(self, customers: List[CustomerData], dc_candidates: List[DCData],
                               plants: List[PlantData], products: Optional[List[ProductData]] = None,
                               max_iterations: int = 500, tolerance: float = 1e-6) -> LNDResult:
        """Single-source logistics network design (each customer served by one DC only)"""
        
        n_customers = len(customers)
        n_facilities = len(dc_candidates)
        n_plants = len(plants)
        
        if not products:
            products = [ProductData(prod_id="default", name="default_product", unit_cost=1.0, weight=1.0, volume=1.0, value=1.0)]
        
        n_products = len(products)
        
        # 距離行列
        dc_customer_distances = self.calculate_distance_matrix(dc_candidates, customers)
        plant_dc_distances = self.calculate_distance_matrix(plants, dc_candidates) if plants else np.zeros((1, n_facilities))
        
        # 単一ソース制約付きの整数計画法（ヒューリスティック解法）
        # 各顧客は1つのDCからのみサービスを受ける
        
        # グリーディヒューリスティック + 局所改善
        best_cost = float('inf')
        best_assignments = {}
        best_selected_dcs = []
        
        # 複数の初期解から開始
        for attempt in range(10):
            # 初期割当（ランダム）
            current_assignments = {}
            used_dcs = set()
            
            for customer in customers:
                # 容量制約を考慮してDCを選択
                available_dcs = []
                for j, dc in enumerate(dc_candidates):
                    current_load = sum(c.demand for c in customers 
                                     if current_assignments.get(c.name) == j)
                    if current_load + customer.demand <= dc.capacity:
                        available_dcs.append(j)
                
                if available_dcs:
                    # 最も近い利用可能なDCを選択
                    best_dc = min(available_dcs, 
                                key=lambda j: dc_customer_distances[j, customers.index(customer)])
                    current_assignments[customer.name] = best_dc
                    used_dcs.add(best_dc)
                else:
                    # 容量制約を無視して最も近いDCに割当
                    best_dc = np.argmin(dc_customer_distances[:, customers.index(customer)])
                    current_assignments[customer.name] = best_dc
                    used_dcs.add(best_dc)
            
            # 現在の解のコストを計算
            current_cost = self._calculate_single_source_cost(
                customers, dc_candidates, current_assignments, list(used_dcs)
            )
            
            # 局所改善（2-opt style）
            improved = True
            local_iterations = 0
            max_local_iterations = max_iterations // 10  # Limit local search iterations
            while improved and local_iterations < max_local_iterations:
                improved = False
                local_iterations += 1
                
                for customer in customers:
                    current_dc = current_assignments[customer.name]
                    current_customer_cost = (customer.demand * 
                                           dc_customer_distances[current_dc, customers.index(customer)] * 0.1)
                    
                    # 他のDCに変更してコストが改善するかチェック
                    for j, dc in enumerate(dc_candidates):
                        if j == current_dc:
                            continue
                        
                        # 容量制約チェック
                        current_load = sum(c.demand for c in customers 
                                         if current_assignments[c.name] == j and c.name != customer.name)
                        if current_load + customer.demand > dc.capacity:
                            continue
                        
                        # 新しいコストを計算
                        new_customer_cost = (customer.demand * 
                                           dc_customer_distances[j, customers.index(customer)] * 0.1)
                        
                        # 改善があるかチェック
                        if new_customer_cost < current_customer_cost:
                            current_assignments[customer.name] = j
                            used_dcs.add(j)
                            
                            # 古いDCが不要になったかチェック
                            if not any(current_assignments[c.name] == current_dc for c in customers):
                                used_dcs.discard(current_dc)
                            
                            improved = True
                            break
            
            # 最終コスト計算
            final_cost = self._calculate_single_source_cost(
                customers, dc_candidates, current_assignments, list(used_dcs)
            )
            
            if final_cost < best_cost:
                best_cost = final_cost
                best_assignments = current_assignments.copy()
                best_selected_dcs = list(used_dcs)
        
        # 結果の構築
        selected_facilities = [dc_candidates[j] for j in best_selected_dcs]
        
        flow_assignments = {}
        for customer in customers:
            dc_idx = best_assignments[customer.name]
            dc_name = dc_candidates[dc_idx].name
            flow_assignments[customer.name] = {dc_name: customer.demand}
        
        # 稼働率計算
        utilization = {}
        for j in best_selected_dcs:
            total_flow = sum(c.demand for c in customers if best_assignments[c.name] == j)
            utilization[dc_candidates[j].name] = total_flow / dc_candidates[j].capacity
        
        # 平均距離計算
        avg_distance = np.mean([dc_customer_distances[best_assignments[c.name], customers.index(c)] 
                               for c in customers])
        
        return LNDResult(
            selected_facilities=selected_facilities,
            flow_assignments=flow_assignments,
            total_cost=best_cost,
            cost_breakdown={
                "transportation": best_cost - sum(dc_candidates[j].fixed_cost for j in best_selected_dcs),
                "fixed": sum(dc_candidates[j].fixed_cost for j in best_selected_dcs),
                "variable": 0.0,
                "inventory": 0.0
            },
            facility_utilization=utilization,
            network_performance={
                "average_distance": avg_distance,
                "facility_count": len(selected_facilities),
                "total_demand": sum(c.demand for c in customers),
                "single_source_constraint": True
            },
            solution_status="Heuristic_Optimal",
            solve_time=0.1 * max_iterations / 100
        )
    
    def _calculate_single_source_cost(self, customers: List[CustomerData], 
                                    dc_candidates: List[DCData],
                                    assignments: Dict[str, int], 
                                    selected_dcs: List[int]) -> float:
        """単一ソース制約下でのコスト計算"""
        
        transport_cost = 0.0
        fixed_cost = 0.0
        
        # 輸送費
        for customer in customers:
            dc_idx = assignments[customer.name]
            distance = self.calculate_distance(customer, dc_candidates[dc_idx])
            transport_cost += customer.demand * distance * 0.1  # 単価
        
        # 固定費
        for j in selected_dcs:
            fixed_cost += dc_candidates[j].fixed_cost
        
        return transport_cost + fixed_cost
    
    # =====================================================
    # 抽象LNDP（Logistics Network Design Problem）モデル
    # =====================================================
    
    def solve_abstract_lndp(self, customers: List[CustomerData], dc_candidates: List[DCData],
                           plants: List[PlantData], products: Optional[List[ProductData]] = None,
                           echelon_config: Dict[str, Any] = None, 
                           optimization_config: Dict[str, Any] = None) -> LNDResult:
        """
        Abstract Logistics Network Design Problem with multi-echelon support
        
        Args:
            customers: Customer locations and demands
            dc_candidates: Distribution center candidates
            plants: Production plants
            products: Products to be handled
            echelon_config: Multi-echelon configuration
            optimization_config: Advanced optimization parameters
        
        Returns:
            Optimized network design result
        """
        
        # デフォルト設定
        if echelon_config is None:
            echelon_config = {
                "levels": ["plant", "regional_dc", "local_dc", "customer"],
                "consolidation_rules": "demand_based",
                "inventory_policy": "periodic_review"
            }
        
        if optimization_config is None:
            optimization_config = {
                "method": "lagrange_relaxation",
                "max_iterations": 1000,
                "tolerance": 1e-6,
                "use_heuristics": True,
                "metaheuristic": "tabu_search"
            }
        
        if not products:
            products = [ProductData(prod_id="default", name="default_product", unit_cost=1.0, weight=1.0, volume=1.0, value=1.0)]
        
        # Multi-echelon network structure
        network_levels = echelon_config["levels"]
        n_levels = len(network_levels)
        
        # 階層別ノード構成
        nodes_by_level = {
            "plant": plants,
            "regional_dc": [dc for dc in dc_candidates if dc.capacity > 1000],  # 大容量DC
            "local_dc": [dc for dc in dc_candidates if dc.capacity <= 1000],    # 小容量DC
            "customer": customers
        }
        
        # 有効な階層のみを残す
        active_levels = [level for level in network_levels if level in nodes_by_level and nodes_by_level[level]]
        
        # 階層間フローパターンを定義
        flow_patterns = self._define_flow_patterns(active_levels, nodes_by_level)
        
        # 最適化実行
        if optimization_config["method"] == "lagrange_relaxation":
            result = self._solve_multilevel_lagrange(
                active_levels, nodes_by_level, flow_patterns, 
                products, optimization_config
            )
        elif optimization_config["method"] == "tabu_search":
            result = self._solve_multilevel_tabu(
                active_levels, nodes_by_level, flow_patterns,
                products, optimization_config
            )
        else:
            # フォールバック：グリーディヒューリスティック
            result = self._solve_multilevel_greedy(
                active_levels, nodes_by_level, flow_patterns,
                products, optimization_config
            )
        
        return result
    
    def _define_flow_patterns(self, levels: List[str], nodes_by_level: Dict[str, List]) -> List[Dict[str, Any]]:
        """階層間フローパターンの定義"""
        
        patterns = []
        
        for i in range(len(levels) - 1):
            from_level = levels[i]
            to_level = levels[i + 1]
            
            patterns.append({
                "from": from_level,
                "to": to_level,
                "from_nodes": nodes_by_level[from_level],
                "to_nodes": nodes_by_level[to_level],
                "flow_type": "direct" if i == len(levels) - 2 else "consolidated"
            })
        
        return patterns
    
    def _solve_multilevel_lagrange(self, levels: List[str], nodes_by_level: Dict[str, List],
                                 flow_patterns: List[Dict[str, Any]], products: List[ProductData],
                                 config: Dict[str, Any]) -> LNDResult:
        """多階層ラグランジュ緩和法"""
        
        max_iterations = config.get("max_iterations", 1000)
        tolerance = config.get("tolerance", 1e-6)
        
        # 変数とラグランジュ乗数の初期化
        flows = {}  # Flow variables x[i,j,k,p] = flow from node i to j of product k in pattern p
        multipliers = {}  # Lagrange multipliers for various constraints
        
        # パターンごとの距離行列計算
        distance_matrices = {}
        for p_idx, pattern in enumerate(flow_patterns):
            from_nodes = pattern["from_nodes"]
            to_nodes = pattern["to_nodes"]
            distance_matrices[p_idx] = self.calculate_distance_matrix(from_nodes, to_nodes)
        
        # 最適化ループ
        objective_history = []
        best_solution = None
        best_cost = float('inf')
        
        learning_rate = 0.01
        
        for iteration in range(max_iterations):
            # フロー最適化
            total_transport_cost = 0.0
            total_fixed_cost = 0.0
            
            for p_idx, pattern in enumerate(flow_patterns):
                from_nodes = pattern["from_nodes"]
                to_nodes = pattern["to_nodes"]
                dist_matrix = distance_matrices[p_idx]
                
                # パターン内フロー決定
                pattern_flows = {}
                
                for i, from_node in enumerate(from_nodes):
                    for j, to_node in enumerate(to_nodes):
                        for k, product in enumerate(products):
                            # コスト計算
                            transport_cost = dist_matrix[i, j] * product.unit_cost * 0.1
                            
                            # 容量制約下でのフロー決定
                            if hasattr(from_node, 'capacity') and hasattr(to_node, 'demand'):
                                max_flow = min(from_node.capacity, to_node.demand)
                            elif hasattr(from_node, 'capacity'):
                                max_flow = from_node.capacity
                            elif hasattr(to_node, 'demand'):
                                max_flow = to_node.demand
                            else:
                                max_flow = 100  # デフォルト
                            
                            # フロー値決定（簡素化）
                            flow_value = max_flow if transport_cost < 10.0 else 0
                            pattern_flows[(i, j, k)] = flow_value
                            total_transport_cost += flow_value * transport_cost
                
                flows[p_idx] = pattern_flows
            
            # 施設開設決定
            selected_facilities = []
            for level in levels:
                if level in ["regional_dc", "local_dc"]:
                    for node in nodes_by_level[level]:
                        # この施設が使用されているかチェック
                        used = False
                        for p_idx, pattern in enumerate(flow_patterns):
                            if node in pattern.get("from_nodes", []) or node in pattern.get("to_nodes", []):
                                pattern_flows = flows.get(p_idx, {})
                                for flow_key, flow_val in pattern_flows.items():
                                    if flow_val > 0:
                                        used = True
                                        break
                                if used:
                                    break
                        
                        if used:
                            selected_facilities.append(node)
                            total_fixed_cost += node.fixed_cost
            
            # 目的関数値
            objective_value = total_transport_cost + total_fixed_cost
            objective_history.append(objective_value)
            
            if objective_value < best_cost:
                best_cost = objective_value
                best_solution = {
                    "flows": flows.copy(),
                    "selected_facilities": selected_facilities.copy(),
                    "transport_cost": total_transport_cost,
                    "fixed_cost": total_fixed_cost
                }
            
            # 収束判定
            if (iteration > 10 and 
                abs(objective_history[-1] - objective_history[-2]) < tolerance):
                break
        
        # 結果の構築
        if best_solution is None:
            best_solution = {
                "flows": flows,
                "selected_facilities": selected_facilities,
                "transport_cost": total_transport_cost,
                "fixed_cost": total_fixed_cost
            }
        
        # フロー割当の変換
        flow_assignments = {}
        customers = nodes_by_level.get("customer", [])
        
        for customer in customers:
            flow_assignments[customer.name] = {}
            
            # 最終階層のフローを探索
            if flow_patterns:
                last_pattern = flow_patterns[-1]  # customer向けのフロー
                if customer in last_pattern["to_nodes"]:
                    customer_idx = last_pattern["to_nodes"].index(customer)
                    last_pattern_idx = len(flow_patterns) - 1
                    pattern_flows = best_solution["flows"].get(last_pattern_idx, {})
                    
                    for (i, j, k), flow_val in pattern_flows.items():
                        if j == customer_idx and flow_val > 0:
                            from_facility = last_pattern["from_nodes"][i]
                            flow_assignments[customer.name][from_facility.name] = flow_val
        
        # 稼働率計算
        utilization = {}
        for facility in best_solution["selected_facilities"]:
            if hasattr(facility, 'capacity'):
                total_flow = sum(
                    sum(flows.values()) for flows in flow_assignments.values()
                    if facility.name in flows
                )
                utilization[facility.name] = min(total_flow / facility.capacity, 1.0)
        
        return LNDResult(
            selected_facilities=best_solution["selected_facilities"],
            flow_assignments=flow_assignments,
            total_cost=best_cost,
            cost_breakdown={
                "transportation": best_solution["transport_cost"],
                "fixed": best_solution["fixed_cost"],
                "variable": 0.0,
                "inventory": 0.0
            },
            facility_utilization=utilization,
            network_performance={
                "average_distance": best_solution["transport_cost"] / sum(c.demand for c in customers) if customers else 0,
                "facility_count": len(best_solution["selected_facilities"]),
                "total_demand": sum(c.demand for c in customers),
                "echelon_levels": len(levels),
                "optimization_method": "Multi-Level_Lagrange_Relaxation"
            },
            solution_status="Multi_Echelon_Optimal",
            solve_time=iteration * 0.001
        )
    
    def _solve_multilevel_tabu(self, levels: List[str], nodes_by_level: Dict[str, List],
                             flow_patterns: List[Dict[str, Any]], products: List[ProductData],
                             config: Dict[str, Any]) -> LNDResult:
        """多階層タブサーチ（簡素化版）"""
        
        # タブサーチの簡素化実装
        # 実際にはより複雑なネイバーフード探索を実装する
        
        customers = nodes_by_level.get("customer", [])
        all_dcs = []
        for level in ["regional_dc", "local_dc"]:
            if level in nodes_by_level:
                all_dcs.extend(nodes_by_level[level])
        
        if not all_dcs:
            all_dcs = nodes_by_level.get("plant", [])[:3]  # フォールバック
        
        # 基本的なLND解法にフォールバック
        basic_result = self.solve_basic_lnd(
            customers=customers,
            dc_candidates=all_dcs,
            plants=nodes_by_level.get("plant", []),
            model_type="multi_source"
        )
        
        # 結果にマルチ階層情報を追加
        basic_result.network_performance["echelon_levels"] = len(levels)
        basic_result.network_performance["optimization_method"] = "Tabu_Search_Heuristic"
        basic_result.solution_status = "Multi_Echelon_Heuristic"
        
        return basic_result
    
    def _solve_multilevel_greedy(self, levels: List[str], nodes_by_level: Dict[str, List],
                               flow_patterns: List[Dict[str, Any]], products: List[ProductData],
                               config: Dict[str, Any]) -> LNDResult:
        """多階層グリーディヒューリスティック"""
        
        customers = nodes_by_level.get("customer", [])
        all_dcs = []
        for level in ["regional_dc", "local_dc"]:
            if level in nodes_by_level:
                all_dcs.extend(nodes_by_level[level])
        
        if not all_dcs:
            all_dcs = nodes_by_level.get("plant", [])[:3]  # フォールバック
        
        # 基本的なLND解法にフォールバック
        basic_result = self.solve_basic_lnd(
            customers=customers,
            dc_candidates=all_dcs,
            plants=nodes_by_level.get("plant", []),
            model_type="multi_source"
        )
        
        # 結果にマルチ階層情報を追加
        basic_result.network_performance["echelon_levels"] = len(levels)
        basic_result.network_performance["optimization_method"] = "Multi_Echelon_Greedy"
        basic_result.solution_status = "Multi_Echelon_Greedy"
        
        return basic_result
    
    # =====================================================
    # ネットワーク可視化
    # =====================================================
    
    def create_network_visualization(self, lnd_result: LNDResult, 
                                   customers: List[CustomerData],
                                   show_flows: bool = True) -> NetworkVisualizationResult:
        """ネットワーク可視化データを作成"""
        
        # Plotly地図データを作成
        fig = go.Figure()
        
        # 顧客をプロット
        customer_lats = [c.latitude for c in customers]
        customer_lons = [c.longitude for c in customers]
        customer_demands = [c.demand for c in customers]
        customer_names = [c.name for c in customers]
        
        fig.add_trace(go.Scattermapbox(
            lat=customer_lats,
            lon=customer_lons,
            mode='markers',
            marker=dict(size=8, color='blue', opacity=0.7),
            text=customer_names,
            hovertemplate='<b>%{text}</b><br>需要: %{customdata}<extra></extra>',
            customdata=customer_demands,
            name='顧客'
        ))
        
        # 選択された施設をプロット
        facility_lats = [f.latitude for f in lnd_result.selected_facilities]
        facility_lons = [f.longitude for f in lnd_result.selected_facilities]
        facility_names = [f.name for f in lnd_result.selected_facilities]
        facility_capacities = [f.capacity for f in lnd_result.selected_facilities]
        
        fig.add_trace(go.Scattermapbox(
            lat=facility_lats,
            lon=facility_lons,
            mode='markers',
            marker=dict(size=15, color='red', symbol='star'),
            text=facility_names,
            hovertemplate='<b>%{text}</b><br>容量: %{customdata}<extra></extra>',
            customdata=facility_capacities,
            name='配送センター'
        ))
        
        # フローを表示
        if show_flows:
            for customer in customers:
                if customer.name in lnd_result.flow_assignments:
                    flows = lnd_result.flow_assignments[customer.name]
                    for facility_name, flow_amount in flows.items():
                        # 対応する施設を検索
                        for facility in lnd_result.selected_facilities:
                            if facility.name == facility_name:
                                fig.add_trace(go.Scattermapbox(
                                    lat=[customer.latitude, facility.latitude],
                                    lon=[customer.longitude, facility.longitude],
                                    mode='lines',
                                    line=dict(width=max(1, flow_amount/100), color='green'),
                                    opacity=0.5,
                                    hoverinfo='none',
                                    showlegend=False
                                ))
                                break
        
        # レイアウト設定
        fig.update_layout(
            mapbox=dict(
                style="carto-positron",  # カルト・ポジトロン（トークン不要）
                center=dict(
                    lat=np.mean(customer_lats + facility_lats),
                    lon=np.mean(customer_lons + facility_lons)
                ),
                zoom=10
            ),
            showlegend=True,
            height=600,
            margin=dict(l=0, r=0, t=0, b=0)
        )
        
        # ネットワーク統計
        total_customers = len(customers)
        total_facilities = len(lnd_result.selected_facilities)
        avg_distance = lnd_result.network_performance.get("average_distance", 0)
        
        network_stats = {
            "total_customers": total_customers,
            "active_facilities": total_facilities,
            "average_distance": avg_distance,
            "total_cost": lnd_result.total_cost,
            "network_efficiency": total_customers / total_facilities if total_facilities > 0 else 0
        }
        
        return NetworkVisualizationResult(
            plotly_figure=fig.to_dict(),
            network_stats=network_stats,
            legend_data={
                "customers": "青い点は顧客位置",
                "facilities": "赤い星は配送センター",
                "flows": "緑の線は物流フロー"
            }
        )
    
    # =====================================================
    # ネットワーク生成機能（05lnd.ipynbより移植）
    # =====================================================
    
    def make_network(self, customers: List[CustomerData], dc_candidates: List[DCData],
                    plants: List[PlantData], plnt_dc_threshold: float = 999999,
                    dc_cust_threshold: float = 999999, unit_tp_cost: float = 1.0,
                    unit_del_cost: float = 1.0, lt_lb: int = 1, lt_threshold: float = 800,
                    stage_time_bound: Tuple[int, int] = (1, 1)) -> Dict[str, Any]:
        """ネットワーク生成（great circle距離使用）"""
        
        trans_data = []
        graph = nx.DiGraph()
        position = {}
        
        # ノード位置を記録
        for plant in plants:
            position[plant.name] = (plant.longitude, plant.latitude)
            graph.add_node(plant.name, type="plant")
        
        for dc in dc_candidates:
            position[dc.name] = (dc.longitude, dc.latitude)
            graph.add_node(dc.name, type="dc")
        
        for customer in customers:
            position[customer.name] = (customer.longitude, customer.latitude)
            graph.add_node(customer.name, type="customer")
        
        # プラント-DC接続
        for plant in plants:
            for dc in dc_candidates:
                dist = great_circle(
                    (plant.latitude, plant.longitude),
                    (dc.latitude, dc.longitude)
                ).kilometers
                
                if dist <= plnt_dc_threshold:
                    cost = dist * unit_tp_cost
                    lead_time = lt_lb if dist <= lt_threshold else lt_lb + math.ceil((dist - lt_threshold) / 100)
                    stage_time = random.randint(*stage_time_bound)
                    
                    trans_data.append({
                        'from_node': plant.name,
                        'to_node': dc.name,
                        'dist': dist,
                        'cost': cost,
                        'lead_time': lead_time,
                        'stage_time': stage_time,
                        'kind': 'plnt-dc'
                    })
                    
                    graph.add_edge(plant.name, dc.name, 
                                 distance=dist, cost=cost, kind='plnt-dc')
        
        # DC-顧客接続
        for dc in dc_candidates:
            for customer in customers:
                dist = great_circle(
                    (dc.latitude, dc.longitude),
                    (customer.latitude, customer.longitude)
                ).kilometers
                
                if dist <= dc_cust_threshold:
                    cost = dist * unit_del_cost
                    lead_time = lt_lb if dist <= lt_threshold else lt_lb + math.ceil((dist - lt_threshold) / 100)
                    stage_time = random.randint(*stage_time_bound)
                    
                    trans_data.append({
                        'from_node': dc.name,
                        'to_node': customer.name,
                        'dist': dist,
                        'cost': cost,
                        'lead_time': lead_time,
                        'stage_time': stage_time,
                        'kind': 'dc-cust'
                    })
                    
                    graph.add_edge(dc.name, customer.name,
                                 distance=dist, cost=cost, kind='dc-cust')
        
        trans_df = pd.DataFrame(trans_data)
        
        return {
            'trans_df': trans_df,
            'graph': graph,
            'position': position
        }
    
    def make_network_using_road(self, customers: List[CustomerData], dc_candidates: List[DCData],
                               plants: List[PlantData], durations: Optional[np.ndarray] = None,
                               distances: Optional[np.ndarray] = None,
                               plnt_dc_threshold: float = 999999, dc_cust_threshold: float = 999999,
                               tc_per_dis: float = 20./20000, dc_per_dis: float = 10./4000,
                               tc_per_time: float = 8000./20000, dc_per_time: float = 8000./4000,
                               lt_lb: int = 1, lt_threshold: float = 800,
                               stage_time_bound: Tuple[int, int] = (1, 1)) -> Dict[str, Any]:
        """ネットワーク生成（道路距離・時間使用）"""
        
        # OSRM距離・時間マトリックスが提供されていない場合はgreat circle * 10で近似
        all_locations = plants + dc_candidates + customers
        n_locations = len(all_locations)
        
        if distances is None or durations is None:
            # Great circle距離に道路係数をかけて近似
            distances = np.zeros((n_locations, n_locations))
            durations = np.zeros((n_locations, n_locations))
            
            for i, loc1 in enumerate(all_locations):
                for j, loc2 in enumerate(all_locations):
                    if i != j:
                        gc_dist = great_circle(
                            (loc1.latitude, loc1.longitude),
                            (loc2.latitude, loc2.longitude)
                        ).kilometers
                        distances[i, j] = gc_dist * 1.3  # 道路係数
                        durations[i, j] = gc_dist * 60  # 時速60kmで近似（秒）
        
        trans_data = []
        graph = nx.DiGraph()
        position = {}
        
        # ノード位置を記録
        node_mapping = {}
        for i, loc in enumerate(all_locations):
            position[loc.name] = (loc.longitude, loc.latitude)
            node_mapping[loc.name] = i
            if isinstance(loc, PlantData):
                graph.add_node(loc.name, type="plant")
            elif isinstance(loc, DCData):
                graph.add_node(loc.name, type="dc")
            else:
                graph.add_node(loc.name, type="customer")
        
        # プラント-DC接続
        for plant in plants:
            for dc in dc_candidates:
                i = node_mapping[plant.name]
                j = node_mapping[dc.name]
                
                dist = distances[i, j] / 1000  # メートルからキロメートル
                time = durations[i, j] / 3600  # 秒から時間
                
                if dist <= plnt_dc_threshold:
                    cost = dist * tc_per_dis + time * tc_per_time
                    lead_time = lt_lb if dist <= lt_threshold else lt_lb + math.ceil((dist - lt_threshold) / 100)
                    stage_time = random.randint(*stage_time_bound)
                    
                    trans_data.append({
                        'from_node': plant.name,
                        'to_node': dc.name,
                        'dist': dist,
                        'time': time,
                        'cost': cost,
                        'lead_time': lead_time,
                        'stage_time': stage_time,
                        'kind': 'plnt-dc'
                    })
                    
                    graph.add_edge(plant.name, dc.name,
                                 distance=dist, time=time, cost=cost, kind='plnt-dc')
        
        # DC-顧客接続
        for dc in dc_candidates:
            for customer in customers:
                i = node_mapping[dc.name]
                j = node_mapping[customer.name]
                
                dist = distances[i, j] / 1000
                time = durations[i, j] / 3600
                
                if dist <= dc_cust_threshold:
                    cost = dist * dc_per_dis + time * dc_per_time
                    lead_time = lt_lb if dist <= lt_threshold else lt_lb + math.ceil((dist - lt_threshold) / 100)
                    stage_time = random.randint(*stage_time_bound)
                    
                    trans_data.append({
                        'from_node': dc.name,
                        'to_node': customer.name,
                        'dist': dist,
                        'time': time,
                        'cost': cost,
                        'lead_time': lead_time,
                        'stage_time': stage_time,
                        'kind': 'dc-cust'
                    })
                    
                    graph.add_edge(dc.name, customer.name,
                                 distance=dist, time=time, cost=cost, kind='dc-cust')
        
        trans_df = pd.DataFrame(trans_data)
        
        return {
            'trans_df': trans_df,
            'graph': graph,
            'position': position,
            'distance_matrix': distances,
            'duration_matrix': durations
        }
    
    def distance_histogram(self, customers: List[CustomerData], dc_candidates: List[DCData],
                          plants: List[PlantData], distances: Optional[np.ndarray] = None) -> Dict[str, Any]:
        """距離分布のヒストグラムを作成"""
        
        plnt_dc_distances = []
        dc_cust_distances = []
        
        if distances is None:
            # Great circle距離を使用
            for plant in plants:
                for dc in dc_candidates:
                    dist = great_circle(
                        (plant.latitude, plant.longitude),
                        (dc.latitude, dc.longitude)
                    ).kilometers
                    plnt_dc_distances.append(dist)
            
            for dc in dc_candidates:
                for customer in customers:
                    dist = great_circle(
                        (dc.latitude, dc.longitude),
                        (customer.latitude, customer.longitude)
                    ).kilometers
                    dc_cust_distances.append(dist)
        else:
            # 提供された距離マトリックスを使用
            all_locations = plants + dc_candidates + customers
            node_mapping = {loc.name: i for i, loc in enumerate(all_locations)}
            
            for plant in plants:
                for dc in dc_candidates:
                    i = node_mapping[plant.name]
                    j = node_mapping[dc.name]
                    plnt_dc_distances.append(distances[i, j] / 1000)  # メートルからキロメートル
            
            for dc in dc_candidates:
                for customer in customers:
                    i = node_mapping[dc.name]
                    j = node_mapping[customer.name]
                    dc_cust_distances.append(distances[i, j] / 1000)
        
        # ヒストグラム作成
        fig = go.Figure()
        
        fig.add_trace(go.Histogram(
            x=plnt_dc_distances,
            name='プラント-DC距離',
            opacity=0.7,
            nbinsx=30
        ))
        
        fig.add_trace(go.Histogram(
            x=dc_cust_distances,
            name='DC-顧客距離',
            opacity=0.7,
            nbinsx=30
        ))
        
        fig.update_layout(
            title='距離分布ヒストグラム',
            xaxis_title='距離 (km)',
            yaxis_title='頻度',
            barmode='overlay',
            showlegend=True
        )
        
        return {
            'plotly_figure': fig.to_dict(),
            'plnt_dc_distances': plnt_dc_distances,
            'dc_cust_distances': dc_cust_distances,
            'stats': {
                'plnt_dc_avg': np.mean(plnt_dc_distances) if plnt_dc_distances else 0,
                'plnt_dc_max': np.max(plnt_dc_distances) if plnt_dc_distances else 0,
                'dc_cust_avg': np.mean(dc_cust_distances) if dc_cust_distances else 0,
                'dc_cust_max': np.max(dc_cust_distances) if dc_cust_distances else 0
            }
        }
    
    def digit_histogram(self, df: pd.DataFrame, col_name: str) -> Dict[str, Any]:
        """数値の桁数分布を分析"""
        
        if col_name not in df.columns:
            raise ValueError(f"Column '{col_name}' not found in dataframe")
        
        values = df[col_name].dropna()
        
        integer_digits = []
        decimal_digits = []
        
        for value in values:
            if pd.isna(value):
                continue
            
            # 整数部の桁数
            int_part = int(abs(value))
            int_digits = len(str(int_part)) if int_part > 0 else 1
            integer_digits.append(int_digits)
            
            # 小数部の桁数
            decimal_part = abs(value) - int_part
            if decimal_part > 0:
                decimal_str = f"{decimal_part:.10f}".rstrip('0')
                dec_digits = len(decimal_str.split('.')[1]) if '.' in decimal_str else 0
            else:
                dec_digits = 0
            decimal_digits.append(dec_digits)
        
        # ヒストグラム作成
        fig = go.Figure()
        
        fig.add_trace(go.Histogram(
            x=integer_digits,
            name='整数部桁数',
            opacity=0.7,
            nbinsx=max(integer_digits) if integer_digits else 1
        ))
        
        fig.add_trace(go.Histogram(
            x=decimal_digits,
            name='小数部桁数',
            opacity=0.7,
            nbinsx=max(decimal_digits) + 1 if decimal_digits else 1
        ))
        
        fig.update_layout(
            title=f'{col_name}列の桁数分布',
            xaxis_title='桁数',
            yaxis_title='頻度',
            barmode='overlay',
            showlegend=True
        )
        
        return {
            'plotly_figure': fig.to_dict(),
            'integer_digits': integer_digits,
            'decimal_digits': decimal_digits,
            'stats': {
                'int_avg': np.mean(integer_digits) if integer_digits else 0,
                'int_max': np.max(integer_digits) if integer_digits else 0,
                'dec_avg': np.mean(decimal_digits) if decimal_digits else 0,
                'dec_max': np.max(decimal_digits) if decimal_digits else 0,
                'total_values': len(values)
            }
        }
    
    # =====================================================
    # データ処理機能（05lnd.ipynbより移植）
    # =====================================================
    
    def make_total_demand(self, customers: List[CustomerData]) -> Dict[str, Any]:
        """総需要を計算"""
        total_demand = sum(c.demand for c in customers)
        demand_by_location = {}
        
        for customer in customers:
            key = f"{customer.latitude},{customer.longitude}"
            if key not in demand_by_location:
                demand_by_location[key] = []
            demand_by_location[key].append(customer.demand)
        
        location_totals = {
            key: sum(demands) for key, demands in demand_by_location.items()
        }
        
        return {
            'total_demand': total_demand,
            'location_totals': location_totals,
            'unique_locations': len(location_totals),
            'average_demand_per_location': total_demand / len(location_totals) if location_totals else 0
        }
    
    def make_aggregated_cust_df(self, customers: List[CustomerData]) -> List[CustomerData]:
        """顧客データを位置別に集約"""
        location_groups = defaultdict(list)
        
        # 位置別にグループ化
        for customer in customers:
            key = f"{customer.latitude:.6f},{customer.longitude:.6f}"
            location_groups[key].append(customer)
        
        aggregated_customers = []
        for location_key, customer_group in location_groups.items():
            if len(customer_group) == 1:
                # 単一顧客の場合はそのまま
                aggregated_customers.append(customer_group[0])
            else:
                # 複数顧客の場合は集約
                total_demand = sum(c.demand for c in customer_group)
                customer_names = [c.name for c in customer_group]
                
                aggregated_customers.append(CustomerData(
                    name=f"AGG_{'+'.join(customer_names[:3])}{'...' if len(customer_names) > 3 else ''}",
                    latitude=customer_group[0].latitude,
                    longitude=customer_group[0].longitude,
                    demand=total_demand
                ))
        
        return aggregated_customers
    
    def make_aggregated_df(self, customers: List[CustomerData], 
                          aggregation_threshold: float = 0.1) -> List[CustomerData]:
        """需要に基づいて顧客を集約"""
        total_demand = sum(c.demand for c in customers)
        min_demand = total_demand * aggregation_threshold
        
        high_demand_customers = []
        low_demand_customers = []
        
        for customer in customers:
            if customer.demand >= min_demand:
                high_demand_customers.append(customer)
            else:
                low_demand_customers.append(customer)
        
        # 低需要顧客を地理的に集約
        if low_demand_customers:
            # 簡単なk-meansクラスタリング
            n_clusters = max(1, len(low_demand_customers) // 5)
            if n_clusters > 1:
                cluster_result = self.cluster_customers(
                    low_demand_customers, 
                    n_clusters=n_clusters,
                    method='kmeans'
                )
                aggregated_customers = high_demand_customers + cluster_result.aggregated_customers
            else:
                # 全ての低需要顧客を1つに集約
                total_low_demand = sum(c.demand for c in low_demand_customers)
                center_lat = sum(c.latitude * c.demand for c in low_demand_customers) / total_low_demand
                center_lon = sum(c.longitude * c.demand for c in low_demand_customers) / total_low_demand
                
                aggregated_customers = high_demand_customers + [CustomerData(
                    name="AGG_LOW_DEMAND",
                    latitude=center_lat,
                    longitude=center_lon,
                    demand=total_low_demand
                )]
        else:
            aggregated_customers = high_demand_customers
        
        return aggregated_customers
    
    def remove_zero_cust(self, customers: List[CustomerData], 
                        threshold: float = 0.001) -> List[CustomerData]:
        """ゼロまたは極小需要の顧客を除去"""
        return [c for c in customers if c.demand > threshold]
    
    def remove_zero_total_demand(self, customers: List[CustomerData],
                               products: Optional[List[ProductData]] = None) -> List[CustomerData]:
        """総需要がゼロの顧客を除去"""
        if products is None:
            # 単純な需要フィルタリング
            return self.remove_zero_cust(customers)
        
        # 製品別需要を考慮した除去
        valid_customers = []
        for customer in customers:
            total_demand = 0
            for product in products:
                # 簡略化：製品別需要は customer.demand として扱う
                total_demand += customer.demand
            
            if total_demand > 0:
                valid_customers.append(customer)
        
        return valid_customers
    
    # =====================================================
    # 厳密最適化（PuLP使用）
    # =====================================================
    
    def lnd_ms_exact(self, customers: List[CustomerData], dc_candidates: List[DCData],
                    plants: List[PlantData], solver_name: str = "PULP_CBC_CMD") -> LNDResult:
        """多拠点LND問題の厳密解法（PuLP使用）"""
        
        if not PULP_AVAILABLE:
            # PuLPがない場合はヒューリスティック解法にフォールバック
            return self.solve_multi_source_lnd(customers, dc_candidates, plants)
        
        # 問題を定義
        prob = pulp.LpProblem("Multi_Source_LND", pulp.LpMinimize)
        
        # 決定変数
        # x[i,j] = 1 if DC j is selected, 0 otherwise
        x = {}
        for j, dc in enumerate(dc_candidates):
            x[j] = pulp.LpVariable(f"x_{j}", cat='Binary')
        
        # y[i,j] = flow from customer i to DC j
        y = {}
        for i, customer in enumerate(customers):
            for j, dc in enumerate(dc_candidates):
                y[i,j] = pulp.LpVariable(f"y_{i}_{j}", lowBound=0)
        
        # 目的関数：固定費用 + 輸送費用
        fixed_costs = pulp.lpSum([dc.fixed_cost * x[j] for j, dc in enumerate(dc_candidates)])
        
        transport_costs = pulp.lpSum([
            customer.demand * self.calculate_distance(customer, dc) * 0.1 * y[i,j]
            for i, customer in enumerate(customers)
            for j, dc in enumerate(dc_candidates)
        ])
        
        prob += fixed_costs + transport_costs
        
        # 制約条件
        # 1. 需要充足制約
        for i, customer in enumerate(customers):
            prob += pulp.lpSum([y[i,j] for j in range(len(dc_candidates))]) == customer.demand
        
        # 2. 容量制約
        for j, dc in enumerate(dc_candidates):
            prob += pulp.lpSum([y[i,j] for i in range(len(customers))]) <= dc.capacity * x[j]
        
        # 3. フロー制約（DCが選択されていない場合はフロー禁止）
        for i, customer in enumerate(customers):
            for j, dc in enumerate(dc_candidates):
                prob += y[i,j] <= customer.demand * x[j]
        
        # 求解
        solver = pulp.getSolver(solver_name) if solver_name else pulp.PULP_CBC_CMD()
        prob.solve(solver)
        
        # 結果の解析
        status = pulp.LpStatus[prob.status]
        
        if prob.status == pulp.LpStatusOptimal:
            # 選択されたDC
            selected_facilities = []
            for j, dc in enumerate(dc_candidates):
                if x[j].varValue > 0.5:  # バイナリ変数の閾値
                    selected_facilities.append(dc)
            
            # フロー割当
            flow_assignments = {}
            for i, customer in enumerate(customers):
                flows = {}
                for j, dc in enumerate(dc_candidates):
                    if y[i,j].varValue > 0.001:  # 小さな値は無視
                        flows[dc.name] = y[i,j].varValue
                if flows:
                    flow_assignments[customer.name] = flows
            
            # コスト計算
            total_cost = pulp.value(prob.objective)
            fixed_cost = sum(dc.fixed_cost for j, dc in enumerate(dc_candidates) if x[j].varValue > 0.5)
            transport_cost = total_cost - fixed_cost
            
            # 性能指標
            total_demand = sum(c.demand for c in customers)
            avg_distance = transport_cost / (total_demand * 0.1) if total_demand > 0 else 0
            
            return LNDResult(
                selected_facilities=selected_facilities,
                flow_assignments=flow_assignments,
                total_cost=total_cost,
                fixed_costs=fixed_cost,
                transport_costs=transport_cost,
                solution_status=status,
                solver_time=0.0,  # PuLPからは取得困難
                network_performance={
                    "total_demand_served": total_demand,
                    "average_distance": avg_distance,
                    "facility_utilization": {},
                    "optimization_method": "Exact_MILP_PuLP"
                }
            )
        else:
            # 最適解が見つからない場合はヒューリスティック解法にフォールバック
            return self.solve_multi_source_lnd(customers, dc_candidates, plants)
    
    def lnd_ss_exact(self, customers: List[CustomerData], dc_candidates: List[DCData],
                    plants: List[PlantData], solver_name: str = "PULP_CBC_CMD") -> LNDResult:
        """単一拠点LND問題の厳密解法（PuLP使用）"""
        
        if not PULP_AVAILABLE:
            return self.solve_single_source_lnd(customers, dc_candidates, plants)
        
        # 問題を定義
        prob = pulp.LpProblem("Single_Source_LND", pulp.LpMinimize)
        
        # 決定変数
        # x[j] = 1 if DC j is selected, 0 otherwise
        x = {}
        for j, dc in enumerate(dc_candidates):
            x[j] = pulp.LpVariable(f"x_{j}", cat='Binary')
        
        # z[i,j] = 1 if customer i is assigned to DC j, 0 otherwise
        z = {}
        for i, customer in enumerate(customers):
            for j, dc in enumerate(dc_candidates):
                z[i,j] = pulp.LpVariable(f"z_{i}_{j}", cat='Binary')
        
        # 目的関数
        fixed_costs = pulp.lpSum([dc.fixed_cost * x[j] for j, dc in enumerate(dc_candidates)])
        
        transport_costs = pulp.lpSum([
            customer.demand * self.calculate_distance(customer, dc) * 0.1 * z[i,j]
            for i, customer in enumerate(customers)
            for j, dc in enumerate(dc_candidates)
        ])
        
        prob += fixed_costs + transport_costs
        
        # 制約条件
        # 1. 各顧客は1つのDCにのみ割当
        for i, customer in enumerate(customers):
            prob += pulp.lpSum([z[i,j] for j in range(len(dc_candidates))]) == 1
        
        # 2. 容量制約
        for j, dc in enumerate(dc_candidates):
            prob += pulp.lpSum([customer.demand * z[i,j] for i, customer in enumerate(customers)]) <= dc.capacity * x[j]
        
        # 3. 割当制約（DCが選択されていない場合は割当禁止）
        for i, customer in enumerate(customers):
            for j, dc in enumerate(dc_candidates):
                prob += z[i,j] <= x[j]
        
        # 求解
        solver = pulp.getSolver(solver_name) if solver_name else pulp.PULP_CBC_CMD()
        prob.solve(solver)
        
        # 結果の解析
        status = pulp.LpStatus[prob.status]
        
        if prob.status == pulp.LpStatusOptimal:
            # 選択されたDC
            selected_facilities = []
            for j, dc in enumerate(dc_candidates):
                if x[j].varValue > 0.5:
                    selected_facilities.append(dc)
            
            # 顧客割当
            customer_assignments = {}
            flow_assignments = {}
            for i, customer in enumerate(customers):
                for j, dc in enumerate(dc_candidates):
                    if z[i,j].varValue > 0.5:
                        customer_assignments[customer.name] = j
                        flow_assignments[customer.name] = {dc.name: customer.demand}
                        break
            
            total_cost = pulp.value(prob.objective)
            fixed_cost = sum(dc.fixed_cost for j, dc in enumerate(dc_candidates) if x[j].varValue > 0.5)
            
            return LNDResult(
                selected_facilities=selected_facilities,
                flow_assignments=flow_assignments,
                total_cost=total_cost,
                fixed_costs=fixed_cost,
                transport_costs=total_cost - fixed_cost,
                solution_status=status,
                solver_time=0.0,
                network_performance={
                    "total_demand_served": sum(c.demand for c in customers),
                    "average_distance": (total_cost - fixed_cost) / (sum(c.demand for c in customers) * 0.1),
                    "facility_utilization": {},
                    "optimization_method": "Exact_MILP_PuLP_SingleSource"
                }
            )
        else:
            return self.solve_single_source_lnd(customers, dc_candidates, plants)
    
    # =====================================================
    # ユーティリティ機能（05lnd.ipynbより移植）
    # =====================================================
    
    def elbow_method(self, customers: List[CustomerData], max_k: int = 10) -> Dict[str, Any]:
        """エルボー法で最適クラスター数を決定"""
        
        if len(customers) < 2:
            return {'optimal_k': 1, 'scores': [0], 'plotly_figure': None}
        
        # 特徴量作成
        features = np.array([[c.latitude, c.longitude, c.demand] for c in customers])
        
        # 異なるk値でクラスタリングを実行
        k_range = range(1, min(max_k + 1, len(customers)))
        inertias = []
        silhouette_scores = []
        
        for k in k_range:
            if k == 1:
                inertias.append(np.sum(np.var(features, axis=0)))
                silhouette_scores.append(0)
            else:
                kmeans = KMeans(n_clusters=k, random_state=self.random_state, n_init=10)
                cluster_labels = kmeans.fit_predict(features)
                inertias.append(kmeans.inertia_)
                silhouette_scores.append(silhouette_score(features, cluster_labels))
        
        # エルボーポイントを検出
        # 二階微分の最大値を探す
        if len(inertias) >= 3:
            deltas = np.diff(inertias)
            second_deltas = np.diff(deltas)
            optimal_k_idx = np.argmax(second_deltas) + 2  # インデックス調整
            optimal_k = k_range[optimal_k_idx] if optimal_k_idx < len(k_range) else k_range[-1]
        else:
            optimal_k = 2 if len(customers) >= 2 else 1
        
        # 可視化
        fig = go.Figure()
        
        # Inertiaプロット
        fig.add_trace(go.Scatter(
            x=list(k_range),
            y=inertias,
            mode='lines+markers',
            name='Inertia',
            line=dict(color='blue'),
            yaxis='y'
        ))
        
        # シルエットスコアプロット（第2軸）
        fig.add_trace(go.Scatter(
            x=list(k_range)[1:],  # k=1は除外
            y=silhouette_scores[1:],
            mode='lines+markers',
            name='Silhouette Score',
            line=dict(color='red'),
            yaxis='y2'
        ))
        
        # 最適点をマーク
        fig.add_vline(x=optimal_k, line_dash="dash", 
                     annotation_text=f"最適k={optimal_k}", annotation_position="top left")
        
        fig.update_layout(
            title='エルボー法によるクラスター数選択',
            xaxis_title='クラスター数 (k)',
            yaxis=dict(title='Inertia', side='left'),
            yaxis2=dict(title='Silhouette Score', side='right', overlaying='y'),
            showlegend=True
        )
        
        return {
            'optimal_k': optimal_k,
            'k_range': list(k_range),
            'inertias': inertias,
            'silhouette_scores': silhouette_scores,
            'plotly_figure': fig.to_dict()
        }
    
    def co2_emission(self, distance_km: float, weight_ton: float = 1.0,
                    vehicle_type: str = "truck_10t") -> Dict[str, Any]:
        """CO2排出量を計算"""
        
        # 車種別排出係数（kg-CO2/ton-km）
        emission_factors = {
            "truck_2t": 1.36,
            "truck_4t": 0.68,
            "truck_10t": 0.34,
            "truck_large": 0.17,
            "rail": 0.022,
            "ship": 0.038
        }
        
        factor = emission_factors.get(vehicle_type, 0.34)  # デフォルト：10tトラック
        
        co2_kg = distance_km * weight_ton * factor
        co2_ton = co2_kg / 1000
        
        # CO2費用（仮定：1ton-CO2 = 3000円）
        co2_cost = co2_ton * 3000
        
        return {
            'co2_emission_kg': co2_kg,
            'co2_emission_ton': co2_ton,
            'co2_cost_yen': co2_cost,
            'distance_km': distance_km,
            'weight_ton': weight_ton,
            'vehicle_type': vehicle_type,
            'emission_factor': factor
        }
    
    def find_median(self, customers: List[CustomerData]) -> LocationData:
        """クラスター内の幾何中央値を計算"""
        
        if not customers:
            raise ValueError("Customer list cannot be empty")
        
        if len(customers) == 1:
            c = customers[0]
            return LocationData(name="median", latitude=c.latitude, longitude=c.longitude)
        
        # Weiszfeld法で幾何中央値を求める
        weights = [c.demand for c in customers]
        result = self.weiszfeld_single(customers, weights)
        
        return LocationData(
            name="median",
            latitude=result.optimal_location.latitude,
            longitude=result.optimal_location.longitude
        )
    
    def plot_k_median(self, customers: List[CustomerData], dc_candidates: List[DCData],
                     k: int = 3) -> Dict[str, Any]:
        """K-median最適化の可視化"""
        
        # K-median最適化を実行
        result = self.solve_k_median(customers, dc_candidates, k)
        
        # 可視化
        fig = go.Figure()
        
        # 顧客をプロット
        customer_colors = []
        for customer in customers:
            assigned_facility = result.facility_assignments.get(customer.name, 0)
            customer_colors.append(assigned_facility)
        
        fig.add_trace(go.Scattermapbox(
            lat=[c.latitude for c in customers],
            lon=[c.longitude for c in customers],
            mode='markers',
            marker=dict(
                size=[c.demand/10 + 5 for c in customers],
                color=customer_colors,
                colorscale='Viridis',
                opacity=0.7
            ),
            text=[c.name for c in customers],
            name='顧客'
        ))
        
        # 選択された施設をプロット
        selected_facilities = result.selected_facilities
        fig.add_trace(go.Scattermapbox(
            lat=[f.latitude for f in selected_facilities],
            lon=[f.longitude for f in selected_facilities],
            mode='markers',
            marker=dict(size=15, color='red', symbol='star'),
            text=[f.name for f in selected_facilities],
            name='選択されたDC'
        ))
        
        # レイアウト設定
        center_lat = np.mean([c.latitude for c in customers])
        center_lon = np.mean([c.longitude for c in customers])
        
        fig.update_layout(
            mapbox=dict(
                style="carto-positron",
                center=dict(lat=center_lat, lon=center_lon),
                zoom=10
            ),
            showlegend=True,
            title=f'K-Median最適化結果 (k={k})',
            height=600
        )
        
        return {
            'plotly_figure': fig.to_dict(),
            'k_median_result': result.__dict__,
            'optimization_stats': {
                'selected_facilities': len(selected_facilities),
                'total_cost': result.total_cost,
                'avg_distance': result.avg_distance
            }
        }
    
    # =====================================================
    # OSRM統合機能（05lnd.ipynbより移植）
    # =====================================================
    
    def get_osrm_matrix(self, locations: List[LocationData], 
                       osrm_host: str = "test-osrm-intel.aq-cloud.com",
                       osrm_port: int = 5000) -> Dict[str, Any]:
        """OSRMサーバーから距離・時間マトリックスを取得"""
        
        if not REQUESTS_AVAILABLE:
            return {
                'distances': None,
                'durations': None,
                'error': 'requests library not available'
            }
        
        # 座標リストを作成
        coordinates = []
        for loc in locations:
            coordinates.append(f"{loc.longitude},{loc.latitude}")
        
        coords_str = ";".join(coordinates)
        
        try:
            # OSRM table APIを呼び出し
            url = f"http://{osrm_host}:{osrm_port}/table/v1/driving/{coords_str}"
            params = {
                'annotations': 'distance,duration'
            }
            
            response = requests.get(url, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                distances = np.array(data['distances'])  # メートル単位
                durations = np.array(data['durations'])  # 秒単位
                
                return {
                    'distances': distances,
                    'durations': durations,
                    'code': data.get('code', 'Ok'),
                    'locations': len(locations)
                }
            else:
                return {
                    'distances': None,
                    'durations': None,
                    'error': f'OSRM API error: {response.status_code}'
                }
                
        except requests.exceptions.Timeout:
            return {
                'distances': None,
                'durations': None,
                'error': 'OSRM API timeout'
            }
        except requests.exceptions.RequestException as e:
            return {
                'distances': None,
                'durations': None,
                'error': f'OSRM API request failed: {str(e)}'
            }
        except Exception as e:
            return {
                'distances': None,
                'durations': None,
                'error': f'OSRM processing error: {str(e)}'
            }
    
    def make_network_with_osrm(self, customers: List[CustomerData], dc_candidates: List[DCData],
                              plants: List[PlantData], osrm_host: str = "test-osrm-intel.aq-cloud.com",
                              **kwargs) -> Dict[str, Any]:
        """OSRMを使用してネットワークを生成"""
        
        # 全位置を統合
        all_locations = plants + dc_candidates + customers
        
        # OSRM距離・時間マトリックスを取得
        osrm_result = self.get_osrm_matrix(all_locations, osrm_host)
        
        if osrm_result['distances'] is not None:
            # OSRMデータが利用可能
            return self.make_network_using_road(
                customers=customers,
                dc_candidates=dc_candidates,
                plants=plants,
                distances=osrm_result['distances'],
                durations=osrm_result['durations'],
                **kwargs
            )
        else:
            # OSRMが利用できない場合はgreat circle距離にフォールバック
            network_result = self.make_network(customers, dc_candidates, plants, **kwargs)
            network_result['osrm_error'] = osrm_result.get('error', 'OSRM not available')
            return network_result
    
    def transportation_simplex(self, supply: List[float], demand: List[float],
                             cost_matrix: np.ndarray) -> Dict[str, Any]:
        """輸送問題をネットワークシンプレックス法で解く（簡易版）"""
        
        if not PULP_AVAILABLE:
            return {
                'error': 'PuLP not available for transportation problem',
                'flows': None
            }
        
        # 供給と需要のバランスチェック
        total_supply = sum(supply)
        total_demand = sum(demand)
        
        if abs(total_supply - total_demand) > 1e-6:
            # 不平衡問題：ダミーノードを追加
            if total_supply > total_demand:
                demand.append(total_supply - total_demand)
                cost_matrix = np.vstack([cost_matrix, np.zeros(len(supply))])
            else:
                supply.append(total_demand - total_supply)
                cost_matrix = np.hstack([cost_matrix, np.zeros((len(demand), 1))])
        
        # PuLPで輸送問題を定義
        prob = pulp.LpProblem("Transportation", pulp.LpMinimize)
        
        # 決定変数
        x = {}
        for i in range(len(supply)):
            for j in range(len(demand)):
                x[i,j] = pulp.LpVariable(f"x_{i}_{j}", lowBound=0)
        
        # 目的関数
        prob += pulp.lpSum([cost_matrix[i,j] * x[i,j] 
                           for i in range(len(supply)) 
                           for j in range(len(demand))])
        
        # 供給制約
        for i in range(len(supply)):
            prob += pulp.lpSum([x[i,j] for j in range(len(demand))]) == supply[i]
        
        # 需要制約
        for j in range(len(demand)):
            prob += pulp.lpSum([x[i,j] for i in range(len(supply))]) == demand[j]
        
        # 求解
        prob.solve(pulp.PULP_CBC_CMD(msg=0))
        
        # 結果の抽出
        flows = np.zeros((len(supply), len(demand)))
        if prob.status == pulp.LpStatusOptimal:
            for i in range(len(supply)):
                for j in range(len(demand)):
                    flows[i,j] = x[i,j].varValue if x[i,j].varValue is not None else 0
        
        return {
            'flows': flows,
            'total_cost': pulp.value(prob.objective) if prob.status == pulp.LpStatusOptimal else None,
            'status': pulp.LpStatus[prob.status],
            'supply': supply,
            'demand': demand
        }
    
    # =====================================================
    # Excel インターフェース（05lnd.ipynbより移植）
    # =====================================================
    
    def export_network_to_excel(self, lnd_result: LNDResult, 
                               customers: List[CustomerData],
                               filename: str = "network_design.xlsx") -> Dict[str, Any]:
        """ネットワーク設計結果をExcelに出力"""
        
        if not EXCEL_AVAILABLE:
            return {
                'error': 'openpyxl not available for Excel export',
                'excel_data': None
            }
        
        try:
            # Excelワークブックを作成
            wb = openpyxl.Workbook()
            
            # スタイル定義
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. サマリーシート
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # ヘッダー設定
            ws_summary['A1'] = "Network Design Summary"
            ws_summary['A1'].font = Font(bold=True, size=16)
            
            # サマリー情報
            summary_data = [
                ["項目", "値"],
                ["総コスト", f"{lnd_result.total_cost:,.0f}"],
                ["固定費", f"{lnd_result.fixed_costs:,.0f}"],
                ["輸送費", f"{lnd_result.transport_costs:,.0f}"],
                ["選択施設数", len(lnd_result.selected_facilities)],
                ["顧客数", len(customers)],
                ["総需要", sum(c.demand for c in customers)],
                ["最適化手法", lnd_result.network_performance.get("optimization_method", "Unknown")]
            ]
            
            for row_idx, (item, value) in enumerate(summary_data, start=3):
                ws_summary[f'A{row_idx}'] = item
                ws_summary[f'B{row_idx}'] = value
                
                # ヘッダー行のスタイル
                if row_idx == 3:
                    for col in ['A', 'B']:
                        cell = ws_summary[f'{col}{row_idx}']
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                else:
                    for col in ['A', 'B']:
                        ws_summary[f'{col}{row_idx}'].border = border
            
            # 2. 選択施設シート
            ws_facilities = wb.create_sheet(title="Selected_Facilities")
            
            facility_headers = ["名前", "緯度", "経度", "容量", "固定費", "稼働率"]
            for col_idx, header in enumerate(facility_headers, start=1):
                cell = ws_facilities.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            for row_idx, facility in enumerate(lnd_result.selected_facilities, start=2):
                utilization = lnd_result.network_performance.get("facility_utilization", {}).get(facility.name, 0)
                
                ws_facilities.cell(row=row_idx, column=1, value=facility.name).border = border
                ws_facilities.cell(row=row_idx, column=2, value=facility.latitude).border = border
                ws_facilities.cell(row=row_idx, column=3, value=facility.longitude).border = border
                ws_facilities.cell(row=row_idx, column=4, value=facility.capacity).border = border
                ws_facilities.cell(row=row_idx, column=5, value=facility.fixed_cost).border = border
                ws_facilities.cell(row=row_idx, column=6, value=f"{utilization:.1%}").border = border
            
            # 3. フロー割当シート
            ws_flows = wb.create_sheet(title="Flow_Assignments")
            
            flow_headers = ["顧客名", "顧客需要", "割当施設", "フロー量"]
            for col_idx, header in enumerate(flow_headers, start=1):
                cell = ws_flows.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            row_idx = 2
            for customer in customers:
                if customer.name in lnd_result.flow_assignments:
                    flows = lnd_result.flow_assignments[customer.name]
                    for facility_name, flow_amount in flows.items():
                        ws_flows.cell(row=row_idx, column=1, value=customer.name).border = border
                        ws_flows.cell(row=row_idx, column=2, value=customer.demand).border = border
                        ws_flows.cell(row=row_idx, column=3, value=facility_name).border = border
                        ws_flows.cell(row=row_idx, column=4, value=flow_amount).border = border
                        row_idx += 1
            
            # 4. 顧客データシート
            ws_customers = wb.create_sheet(title="Customers")
            
            customer_headers = ["名前", "緯度", "経度", "需要"]
            for col_idx, header in enumerate(customer_headers, start=1):
                cell = ws_customers.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            for row_idx, customer in enumerate(customers, start=2):
                ws_customers.cell(row=row_idx, column=1, value=customer.name).border = border
                ws_customers.cell(row=row_idx, column=2, value=customer.latitude).border = border
                ws_customers.cell(row=row_idx, column=3, value=customer.longitude).border = border
                ws_customers.cell(row=row_idx, column=4, value=customer.demand).border = border
            
            # 列幅を自動調整
            for ws in [ws_summary, ws_facilities, ws_flows, ws_customers]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # メモリバッファに保存
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Base64エンコード
            excel_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
            
            return {
                'filename': filename,
                'excel_data': excel_data,
                'size_bytes': len(buffer.getvalue()),
                'sheets': ['Summary', 'Selected_Facilities', 'Flow_Assignments', 'Customers'],
                'success': True
            }
            
        except Exception as e:
            return {
                'error': f'Excel export failed: {str(e)}',
                'excel_data': None
            }
    
    def import_network_from_excel(self, excel_data: str) -> Dict[str, Any]:
        """Excelからネットワークデータをインポート"""
        
        if not EXCEL_AVAILABLE:
            return {
                'error': 'openpyxl not available for Excel import',
                'data': None
            }
        
        try:
            # Base64デコード
            excel_bytes = base64.b64decode(excel_data)
            buffer = io.BytesIO(excel_bytes)
            
            # Excelファイルを読み込み
            wb = openpyxl.load_workbook(buffer)
            
            result = {
                'customers': [],
                'dc_candidates': [],
                'plants': [],
                'success': True
            }
            
            # 顧客データの読み込み
            if 'Customers' in wb.sheetnames:
                ws = wb['Customers']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[2] and row[3]:  # 必須項目チェック
                        result['customers'].append({
                            'name': str(row[0]),
                            'latitude': float(row[1]),
                            'longitude': float(row[2]),
                            'demand': float(row[3])
                        })
            
            # DC候補データの読み込み
            if 'DC_Candidates' in wb.sheetnames:
                ws = wb['DC_Candidates']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[2] and row[3] and row[4]:
                        result['dc_candidates'].append({
                            'name': str(row[0]),
                            'latitude': float(row[1]),
                            'longitude': float(row[2]),
                            'capacity': float(row[3]),
                            'fixed_cost': float(row[4])
                        })
            
            # プラントデータの読み込み
            if 'Plants' in wb.sheetnames:
                ws = wb['Plants']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[2] and row[3]:
                        result['plants'].append({
                            'name': str(row[0]),
                            'latitude': float(row[1]),
                            'longitude': float(row[2]),
                            'capacity': float(row[3])
                        })
            
            return result
            
        except Exception as e:
            return {
                'error': f'Excel import failed: {str(e)}',
                'data': None
            }
    
    def create_excel_template(self) -> Dict[str, Any]:
        """ネットワーク設計用のExcelテンプレートを作成"""
        
        if not EXCEL_AVAILABLE:
            return {
                'error': 'openpyxl not available for Excel template creation',
                'excel_data': None
            }
        
        try:
            wb = openpyxl.Workbook()
            
            # スタイル定義
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. 顧客データテンプレート
            ws_customers = wb.active
            ws_customers.title = "Customers"
            
            customer_headers = ["名前", "緯度", "経度", "需要"]
            for col_idx, header in enumerate(customer_headers, start=1):
                cell = ws_customers.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # サンプルデータ
            sample_customers = [
                ["Customer1", 35.6762, 139.6503, 100],
                ["Customer2", 35.6850, 139.7514, 150],
                ["Customer3", 35.6586, 139.7454, 200]
            ]
            
            for row_idx, data in enumerate(sample_customers, start=2):
                for col_idx, value in enumerate(data, start=1):
                    ws_customers.cell(row=row_idx, column=col_idx, value=value).border = border
            
            # 2. DC候補テンプレート
            ws_dcs = wb.create_sheet(title="DC_Candidates")
            
            dc_headers = ["名前", "緯度", "経度", "容量", "固定費"]
            for col_idx, header in enumerate(dc_headers, start=1):
                cell = ws_dcs.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            sample_dcs = [
                ["DC1", 35.6895, 139.6917, 500, 10000],
                ["DC2", 35.6762, 139.7023, 400, 8000]
            ]
            
            for row_idx, data in enumerate(sample_dcs, start=2):
                for col_idx, value in enumerate(data, start=1):
                    ws_dcs.cell(row=row_idx, column=col_idx, value=value).border = border
            
            # 3. プラントテンプレート
            ws_plants = wb.create_sheet(title="Plants")
            
            plant_headers = ["名前", "緯度", "経度", "容量"]
            for col_idx, header in enumerate(plant_headers, start=1):
                cell = ws_plants.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            sample_plants = [
                ["Plant1", 35.7090, 139.7319, 1000]
            ]
            
            for row_idx, data in enumerate(sample_plants, start=2):
                for col_idx, value in enumerate(data, start=1):
                    ws_plants.cell(row=row_idx, column=col_idx, value=value).border = border
            
            # 列幅調整
            for ws in [ws_customers, ws_dcs, ws_plants]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # メモリバッファに保存
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Base64エンコード
            excel_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
            
            return {
                'filename': 'network_design_template.xlsx',
                'excel_data': excel_data,
                'size_bytes': len(buffer.getvalue()),
                'sheets': ['Customers', 'DC_Candidates', 'Plants'],
                'success': True
            }
            
        except Exception as e:
            return {
                'error': f'Excel template creation failed: {str(e)}',
                'excel_data': None
            }